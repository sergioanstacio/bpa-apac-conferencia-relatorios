import collections
import zipfile
import os
import sys
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─────────────────────────────────────────────────────────────
# 1. SIGTAP
# ─────────────────────────────────────────────────────────────

SIGTAP_URL   = "https://github.com/RenatoKR/SIGTAP/raw/main/tabelas/TabelaUnificada_202605_v2605210940.zip"
SIGTAP_CACHE = os.path.join(os.path.dirname(__file__), "sigtap_cache.zip")

EXT_MES = {
    'JAN': 'Janeiro', 'FEV': 'Fevereiro', 'MAR': 'Março',
    'ABR': 'Abril',   'MAI': 'Maio',      'JUN': 'Junho',
    'JUL': 'Julho',   'AGO': 'Agosto',    'SET': 'Setembro',
    'OUT': 'Outubro', 'NOV': 'Novembro',  'DEZ': 'Dezembro',
}
EXTENSOES_VALIDAS = set(EXT_MES.keys()) | {'TXT'}


def validar_extensao(caminho):
    ext = os.path.splitext(caminho)[1].upper().lstrip('.')
    return ext in EXTENSOES_VALIDAS


def baixar_sigtap(callback=None):
    if os.path.exists(SIGTAP_CACHE):
        return SIGTAP_CACHE
    if callback:
        callback("Baixando tabela SIGTAP do DATASUS…")
    urllib.request.urlretrieve(SIGTAP_URL, SIGTAP_CACHE)
    return SIGTAP_CACHE


def carregar_procedimentos(sigtap_zip):
    """
    Carrega tabela SIGTAP.
    Para procedimentos ambulatoriais (BPA/APAC), o valor de referência é o
    VL_SA (Valor Ambulatorial). Se VL_SA for zero, usa VL_SH + VL_SP
    (procedimentos hospitalares/profissionais).
    """
    proc_data = {}
    with zipfile.ZipFile(sigtap_zip) as z:
        with z.open("tb_procedimento.txt") as f:
            for line in f:
                line = line.decode("latin1").rstrip("\r\n")
                if len(line) >= 318:
                    code = line[0:10].strip()
                    name = line[10:260].strip()
                    try:
                        vl_sh = int(line[282:294].strip())
                        vl_sa = int(line[294:306].strip())
                        vl_sp = int(line[306:318].strip())
                    except Exception:
                        vl_sh = vl_sa = vl_sp = 0

                    if vl_sa > 0:
                        val = vl_sa / 100
                    else:
                        val = (vl_sh + vl_sp) / 100

                    proc_data[code] = {"name": name, "value": val}
    return proc_data


# ─────────────────────────────────────────────────────────────
# 2. Leitura do arquivo
#    BPA-C: linhas "02"  →  sem paciente, sem data
#    BPA-I: linhas "03"  →  com paciente e data
# ─────────────────────────────────────────────────────────────

def ler_arquivo(caminho):
    """
    Retorna:
      info   : dict  {cnes, nome_cnes, competencia}
      bpa_i  : {nome_paciente: {data: {proc: qtd}}}
      bpa_c  : {cbo: {proc: qtd}}
      all_dates_i: set de datas (BPA-I)
      has_i, has_c : bool
    """
    info  = {"cnes": "", "nome_cnes": "", "competencia": ""}
    bpa_i = collections.defaultdict(
        lambda: collections.defaultdict(lambda: collections.defaultdict(int)))
    bpa_c = collections.defaultdict(lambda: collections.defaultdict(int))
    all_dates_i = set()
    cbos = set()

    with open(caminho, encoding="ISO-8859-1") as f:
        for line in f:
            line = line.rstrip("\r\n")

            # ── Cabeçalho ──
            if line.startswith("01"):
                info["nome_cnes"]   = line[29:59].strip()
                comp_raw            = line[7:13]
                info["competencia"] = f"{comp_raw[4:6]}/{comp_raw[0:4]}"
                continue

            # ── BPA-C (02) ──
            # pos: 1-2 ident | 3-9 cnes | 10-15 cmp | 16-21 cbo | 22-24 flh
            #      25-26 seq | 27-36 pa  | 37-39 idade | 40-45 qt | 46-48 org
            if line.startswith("02") and len(line) >= 45:
                if not info["cnes"]:
                    info["cnes"] = line[2:9]
                cbo  = line[15:21].strip()
                proc = line[26:36].strip()
                try:
                    qtd = int(line[39:45].strip())
                except ValueError:
                    qtd = 0
                if proc:
                    bpa_c[cbo][proc] += qtd

            # ── BPA-I (03) ──
            # pos: 1-2 ident | 3-9 cnes | 10-15 cmp | 16-30 cns_med | 31-36 cbo
            #      37-44 dtaten | 45-47 flh | 48-49 seq | 50-59 pa | 60-74 cns_pac
            #      75 sexo | 76-81 ibge | 82-85 cid | 86-88 idade | 89-94 qt
            #      ... | 113-142 nome_pac
            elif line.startswith("03") and len(line) >= 142:
                if not info["cnes"]:
                    info["cnes"] = line[2:9]
                nome     = line[112:142].strip()
                proc     = line[49:59].strip()
                data_raw = line[36:44].strip()
                try:
                    qtd = int(line[88:94].strip())
                except ValueError:
                    qtd = 0
                data = (f"{data_raw[6:8]}/{data_raw[4:6]}/{data_raw[0:4]}"
                        if len(data_raw) == 8 else "Sem data")
                if nome:
                    bpa_i[nome][data][proc] += qtd
                    all_dates_i.add(data)

    has_i = bool(bpa_i)
    has_c = bool(bpa_c)
    return info, bpa_i, bpa_c, all_dates_i, has_i, has_c


# ─────────────────────────────────────────────────────────────
# 3. Estilos compartilhados
# ─────────────────────────────────────────────────────────────

def _estilos():
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(c): return PatternFill("solid", start_color=c)
    def font(bold=False, size=10, color="000000"):
        return Font(name="Arial", bold=bold, size=size, color=color)

    return {
        "border":       border,
        "fill_hdr_bg":  fill("0D2E4E"),
        "fill_hdr_lbl": fill("1F4E79"),
        "fill_hdr_val": fill("2E75B6"),
        "fill_col_hdr": fill("1F4E79"),
        "fill_name":    fill("2E75B6"),
        "fill_total":   fill("1F4E79"),
        "fill_date":    fill("BDD7EE"),
        "fill_even":    fill("DEEAF1"),
        "fill_odd":     fill("FFFFFF"),
        "fill_cbo":     fill("375623"),   # verde escuro para CBO
        "fill_cbo_lgt": fill("E2EFDA"),   # verde claro fundo proc
        "fill_cbo_even":fill("D9EAD3"),
        "font_title":   font(bold=True, size=13, color="FFFFFF"),
        "font_lbl":     font(bold=True, size=10, color="BDD7EE"),
        "font_val":     font(bold=True, size=10, color="FFFFFF"),
        "font_col_hdr": font(bold=True, size=10, color="FFFFFF"),
        "font_name":    font(bold=True, size=10, color="FFFFFF"),
        "font_total":   font(bold=True, size=10, color="FFFFFF"),
        "font_date":    font(bold=True, size=10, color="1F4E79"),
        "font_cbo":     font(bold=True, size=10, color="FFFFFF"),
        "font_proc":    font(size=10),
        "al_center":    Alignment(horizontal="center", vertical="center"),
        "al_left":      Alignment(horizontal="left",   vertical="center"),
        "al_right":     Alignment(horizontal="right",  vertical="center"),
    }


def _cabecalho_sheet(ws, titulo, cnes, nome_cnes, competencia, ncols=4):
    s = _estilos()
    last = chr(ord('A') + ncols - 1)

    ws.merge_cells(f"A1:{last}1")
    ws["A1"].value     = titulo
    ws["A1"].font      = s["font_title"]
    ws["A1"].fill      = s["fill_hdr_bg"]
    ws["A1"].alignment = s["al_center"]
    ws.row_dimensions[1].height = 26

    for row_n, label, value in [
        (2, "CNES", cnes),
        (3, "NOME DO ESTABELECIMENTO", nome_cnes),
        (4, "COMPETÊNCIA", competencia),
    ]:
        ws[f"A{row_n}"].value     = label
        ws[f"A{row_n}"].font      = s["font_lbl"]
        ws[f"A{row_n}"].fill      = s["fill_hdr_lbl"]
        ws[f"A{row_n}"].alignment = s["al_center"]
        ws.merge_cells(f"B{row_n}:{last}{row_n}")
        ws[f"B{row_n}"].value     = value
        ws[f"B{row_n}"].font      = s["font_val"]
        ws[f"B{row_n}"].fill      = s["fill_hdr_val"]
        ws[f"B{row_n}"].alignment = s["al_left"]
        ws.row_dimensions[row_n].height = 18

    ws.merge_cells(f"A5:{last}5")
    ws["A5"].fill = s["fill_hdr_bg"]
    ws.row_dimensions[5].height = 6


# ─────────────────────────────────────────────────────────────
# 4. BPA-I: relatório por paciente → data → procedimento
# ─────────────────────────────────────────────────────────────

def build_sheet_bpa_i(ws, title_suffix, patients_data, cnes, nome_cnes, competencia, proc_data):
    s = _estilos()
    _cabecalho_sheet(ws, f"RELATÓRIO BPA-I — INDIVIDUALIZADO{title_suffix}",
                     cnes, nome_cnes, competencia, ncols=4)

    for col, title in [
        ("A6", "Cód. Procedimento"),
        ("B6", "Descrição do Procedimento"),
        ("C6", "Quantidade"),
        ("D6", "Valor Unit. (R$)"),
    ]:
        ws[col].value     = title
        ws[col].font      = s["font_col_hdr"]
        ws[col].fill      = s["fill_col_hdr"]
        ws[col].alignment = s["al_center"]
        ws[col].border    = s["border"]
    ws.row_dimensions[6].height = 20
    ws.freeze_panes = "A7"

    row_idx = 7
    for nome, datas in patients_data:
        total_qtd = sum(q for dp in datas.values() for q in dp.values())
        total_val = sum(q * proc_data.get(p, {"value": 0})["value"]
                        for dp in datas.values() for p, q in dp.items())

        ws.merge_cells(f"A{row_idx}:D{row_idx}")
        c = ws[f"A{row_idx}"]
        c.value = nome; c.font = s["font_name"]; c.fill = s["fill_name"]
        c.alignment = s["al_left"]; c.border = s["border"]
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

        for data in sorted(datas.keys(), key=lambda d: (d[6:10], d[3:5], d[0:2])):
            ws.merge_cells(f"A{row_idx}:D{row_idx}")
            c = ws[f"A{row_idx}"]
            c.value = f"  📅 {data}"; c.font = s["font_date"]; c.fill = s["fill_date"]
            c.alignment = s["al_left"]; c.border = s["border"]
            ws.row_dimensions[row_idx].height = 17
            row_idx += 1

            for i, (proc, qtd) in enumerate(sorted(datas[data].items())):
                fill  = s["fill_even"] if i % 2 == 0 else s["fill_odd"]
                pinfo = proc_data.get(proc, {"name": "", "value": 0.0})
                ws[f"A{row_idx}"].value = f"  {proc}"
                ws[f"B{row_idx}"].value = f"  {pinfo['name']}"
                ws[f"C{row_idx}"].value = qtd
                ws[f"D{row_idx}"].value = pinfo["value"]
                ws[f"D{row_idx}"].number_format = 'R$ #,##0.00'
                for col in ["A", "B", "C", "D"]:
                    ws[f"{col}{row_idx}"].font   = s["font_proc"]
                    ws[f"{col}{row_idx}"].fill   = fill
                    ws[f"{col}{row_idx}"].border = s["border"]
                    ws[f"{col}{row_idx}"].alignment = (
                        s["al_center"] if col in ["A", "C", "D"] else s["al_left"])
                ws.row_dimensions[row_idx].height = 16
                row_idx += 1

        # Total paciente
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        ws[f"A{row_idx}"].value = "TOTAL DO PACIENTE"
        ws[f"A{row_idx}"].font  = s["font_total"]; ws[f"A{row_idx}"].fill = s["fill_total"]
        ws[f"A{row_idx}"].alignment = s["al_right"]; ws[f"A{row_idx}"].border = s["border"]
        ws[f"C{row_idx}"].value = total_qtd
        ws[f"C{row_idx}"].font  = s["font_total"]; ws[f"C{row_idx}"].fill = s["fill_total"]
        ws[f"C{row_idx}"].alignment = s["al_center"]; ws[f"C{row_idx}"].border = s["border"]
        ws[f"D{row_idx}"].value = total_val
        ws[f"D{row_idx}"].font  = s["font_total"]; ws[f"D{row_idx}"].fill = s["fill_total"]
        ws[f"D{row_idx}"].alignment = s["al_right"]; ws[f"D{row_idx}"].border = s["border"]
        ws[f"D{row_idx}"].number_format = 'R$ #,##0.00'
        ws.row_dimensions[row_idx].height = 17
        row_idx += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18


# ─────────────────────────────────────────────────────────────
# 5. BPA-C: relatório por CBO → procedimento → quantidade
# ─────────────────────────────────────────────────────────────

def build_sheet_bpa_c(ws, cnes, nome_cnes, competencia, bpa_c, proc_data, cbo_names):
    s = _estilos()
    _cabecalho_sheet(ws, "RELATÓRIO BPA-C — CONSOLIDADO",
                     cnes, nome_cnes, competencia, ncols=4)

    for col, title in [
        ("A6", "Cód. Procedimento"),
        ("B6", "Descrição do Procedimento"),
        ("C6", "Quantidade"),
        ("D6", "Valor Unit. (R$)"),
    ]:
        ws[col].value     = title
        ws[col].font      = s["font_col_hdr"]
        ws[col].fill      = s["fill_col_hdr"]
        ws[col].alignment = s["al_center"]
        ws[col].border    = s["border"]
    ws.row_dimensions[6].height = 20
    ws.freeze_panes = "A7"

    # Totais gerais
    grand_qtd = 0
    grand_val = 0.0

    row_idx = 7
    for cbo in sorted(bpa_c.keys()):
        procs = bpa_c[cbo]
        cbo_label = cbo_names.get(cbo, "")
        cbo_header = f"  CBO {cbo}" + (f"  —  {cbo_label}" if cbo_label else "")

        total_qtd_cbo = sum(procs.values())
        total_val_cbo = sum(q * proc_data.get(p, {"value": 0})["value"]
                            for p, q in procs.items())
        grand_qtd += total_qtd_cbo
        grand_val += total_val_cbo

        # Linha CBO
        ws.merge_cells(f"A{row_idx}:D{row_idx}")
        c = ws[f"A{row_idx}"]
        c.value = cbo_header; c.font = s["font_cbo"]; c.fill = s["fill_cbo"]
        c.alignment = s["al_left"]; c.border = s["border"]
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

        for i, (proc, qtd) in enumerate(sorted(procs.items())):
            fill  = s["fill_cbo_even"] if i % 2 == 0 else s["fill_odd"]
            pinfo = proc_data.get(proc, {"name": "", "value": 0.0})
            ws[f"A{row_idx}"].value = f"  {proc}"
            ws[f"B{row_idx}"].value = f"  {pinfo['name']}"
            ws[f"C{row_idx}"].value = qtd
            ws[f"D{row_idx}"].value = pinfo["value"]
            ws[f"D{row_idx}"].number_format = 'R$ #,##0.00'
            for col in ["A", "B", "C", "D"]:
                ws[f"{col}{row_idx}"].font   = s["font_proc"]
                ws[f"{col}{row_idx}"].fill   = fill
                ws[f"{col}{row_idx}"].border = s["border"]
                ws[f"{col}{row_idx}"].alignment = (
                    s["al_center"] if col in ["A", "C", "D"] else s["al_left"])
            ws.row_dimensions[row_idx].height = 16
            row_idx += 1

        # Total CBO
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        ws[f"A{row_idx}"].value = f"TOTAL CBO {cbo}"
        ws[f"A{row_idx}"].font  = s["font_total"]; ws[f"A{row_idx}"].fill = s["fill_total"]
        ws[f"A{row_idx}"].alignment = s["al_right"]; ws[f"A{row_idx}"].border = s["border"]
        ws[f"C{row_idx}"].value = total_qtd_cbo
        ws[f"C{row_idx}"].font  = s["font_total"]; ws[f"C{row_idx}"].fill = s["fill_total"]
        ws[f"C{row_idx}"].alignment = s["al_center"]; ws[f"C{row_idx}"].border = s["border"]
        ws[f"D{row_idx}"].value = total_val_cbo
        ws[f"D{row_idx}"].font  = s["font_total"]; ws[f"D{row_idx}"].fill = s["fill_total"]
        ws[f"D{row_idx}"].alignment = s["al_right"]; ws[f"D{row_idx}"].border = s["border"]
        ws[f"D{row_idx}"].number_format = 'R$ #,##0.00'
        ws.row_dimensions[row_idx].height = 17
        row_idx += 1

    # Total Geral
    fill_geral = PatternFill("solid", start_color="0D2E4E")
    font_geral = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.merge_cells(f"A{row_idx}:B{row_idx}")
    ws[f"A{row_idx}"].value = "TOTAL GERAL"
    ws[f"A{row_idx}"].font  = font_geral; ws[f"A{row_idx}"].fill = fill_geral
    ws[f"A{row_idx}"].alignment = s["al_right"]; ws[f"A{row_idx}"].border = s["border"]
    ws[f"C{row_idx}"].value = grand_qtd
    ws[f"C{row_idx}"].font  = font_geral; ws[f"C{row_idx}"].fill = fill_geral
    ws[f"C{row_idx}"].alignment = s["al_center"]; ws[f"C{row_idx}"].border = s["border"]
    ws[f"D{row_idx}"].value = grand_val
    ws[f"D{row_idx}"].font  = font_geral; ws[f"D{row_idx}"].fill = fill_geral
    ws[f"D{row_idx}"].alignment = s["al_right"]; ws[f"D{row_idx}"].border = s["border"]
    ws[f"D{row_idx}"].number_format = 'R$ #,##0.00'
    ws.row_dimensions[row_idx].height = 20

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18


# ─────────────────────────────────────────────────────────────
# 6. Carrega nomes de CBO do SIGTAP
# ─────────────────────────────────────────────────────────────

def carregar_cbos(sigtap_zip):
    cbo_names = {}
    try:
        with zipfile.ZipFile(sigtap_zip) as z:
            # Tenta tb_ocupacao.txt (layout: CO_OCUPACAO 1-6, NO_OCUPACAO 7-...)
            names_in_zip = z.namelist()
            cbo_file = next((n for n in names_in_zip if "ocupacao" in n.lower()), None)
            if cbo_file:
                with z.open(cbo_file) as f:
                    for line in f:
                        line = line.decode("latin1").rstrip("\r\n")
                        if len(line) >= 7:
                            code = line[0:6].strip()
                            name = line[6:].strip()
                            cbo_names[code] = name
    except Exception:
        pass
    return cbo_names


# ─────────────────────────────────────────────────────────────
# 7. Função principal
# ─────────────────────────────────────────────────────────────

def gerar_excel(caminho_mai, caminho_saida=None, callback=None):
    ext = os.path.splitext(caminho_mai)[1].upper().lstrip('.')
    if ext not in EXTENSOES_VALIDAS:
        exts = ', '.join(sorted(EXTENSOES_VALIDAS))
        raise ValueError(f"Extensao '.{ext}' nao suportada. Aceitas: {exts}")

    if callback:
        mes = EXT_MES.get(ext, "")
        callback(f"Carregando tabela SIGTAP{'  (' + mes + ')' if mes else ''}…")

    sigtap_zip = baixar_sigtap(callback)
    proc_data  = carregar_procedimentos(sigtap_zip)
    cbo_names  = carregar_cbos(sigtap_zip)

    if callback:
        callback("Lendo arquivo de produção…")

    info, bpa_i, bpa_c, all_dates_i, has_i, has_c = ler_arquivo(caminho_mai)
    cnes        = info["cnes"]
    nome_cnes   = info["nome_cnes"]
    competencia = info["competencia"]

    sorted_patients = sorted(bpa_i.items(), key=lambda x: x[0])
    sorted_dates    = sorted(all_dates_i, key=lambda d: (d[6:10], d[3:5], d[0:2]))

    tipos = []
    if has_i: tipos.append("BPA-I")
    if has_c: tipos.append("BPA-C")
    if callback:
        callback(f"Gerando relatório {'/'.join(tipos)} — "
                 f"{len(sorted_patients)} pacientes, {len(sorted_dates)} datas…")

    wb = Workbook()
    first = True

    # ── Abas BPA-I ──────────────────────────────
    if has_i:
        ws_all = wb.active if first else wb.create_sheet()
        first  = False
        ws_all.title = "BPA-I Todos"
        build_sheet_bpa_i(ws_all, "", sorted_patients,
                          cnes, nome_cnes, competencia, proc_data)

        for data in sorted_dates:
            filtered = [(n, {data: d[data]}) for n, d in sorted_patients if data in d]
            if not filtered:
                continue
            ws_d = wb.create_sheet(title=data.replace("/", "-"))
            build_sheet_bpa_i(ws_d, f" — {data}", filtered,
                              cnes, nome_cnes, competencia, proc_data)

    # ── Aba BPA-C ───────────────────────────────
    if has_c:
        ws_c = wb.active if first else wb.create_sheet(title="BPA-C Consolidado")
        first = False
        if ws_c.title == "Sheet":
            ws_c.title = "BPA-C Consolidado"
        build_sheet_bpa_c(ws_c, cnes, nome_cnes, competencia,
                          bpa_c, proc_data, cbo_names)

    # ── Aba Resumo Geral (sempre a primeira) ─────────────────
    if callback:
        callback("Gerando Resumo Geral…")
    ws_res = wb.create_sheet(title="Resumo Geral", index=0)
    build_sheet_resumo(ws_res, cnes, nome_cnes, competencia,
                       bpa_i, bpa_c, proc_data, sorted_dates)
    wb.active = ws_res  # abrir no resumo por padrão

    if caminho_saida is None:
        base = os.path.splitext(caminho_mai)[0]
        caminho_saida = base + "_relatorio.xlsx"

    wb.save(caminho_saida)
    return caminho_saida, len(sorted_patients), len(sorted_dates), has_i, has_c


# ─────────────────────────────────────────────────────────────
# 8. CLI
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python gerar_relatorio.py <arquivo.MAI> [saida.xlsx]")
        sys.exit(1)

    entrada = sys.argv[1]
    saida   = sys.argv[2] if len(sys.argv) > 2 else None

    saida, n_pac, n_dat, has_i, has_c = gerar_excel(
        entrada, saida, callback=lambda msg: print(f"  → {msg}")
    )
    tipos = []
    if has_i: tipos.append(f"BPA-I ({n_pac} pacientes, {n_dat} datas)")
    if has_c: tipos.append("BPA-C (consolidado)")
    print(f"\n Relatório salvo em: {saida}")
    print(f"   {' | '.join(tipos)}")


# ─────────────────────────────────────────────────────────────
# 9. Aba Resumo Geral — todos os procedimentos (BPA-I + BPA-C)
# ─────────────────────────────────────────────────────────────

def build_sheet_resumo(ws, cnes, nome_cnes, competencia,
                       bpa_i, bpa_c, proc_data, sorted_dates):
    s = _estilos()
    _cabecalho_sheet(ws, "RESUMO GERAL DE PRODUÇÃO",
                     cnes, nome_cnes, competencia, ncols=5)
    for col, title in [
        ("A6", "Cód. Procedimento"),
        ("B6", "Descrição do Procedimento"),
        ("C6", "Quantidade"),
        ("D6", "Valor Unit. (R$)"),
        ("E6", "Valor Total (R$)"),
    ]:
        ws[col].value     = title
        ws[col].font      = s["font_col_hdr"]
        ws[col].fill      = s["fill_col_hdr"]
        ws[col].alignment = s["al_center"]
        ws[col].border    = s["border"]
    ws.row_dimensions[6].height = 20
    ws.freeze_panes = "A7"

    # Agregar todos os procedimentos (BPA-I + BPA-C)
    proc_totals = collections.defaultdict(int)
    for d in bpa_i.values():
        for dp in d.values():
            for p, q in dp.items():
                proc_totals[p] += q
    for procs in bpa_c.values():
        for p, q in procs.items():
            proc_totals[p] += q

    sorted_procs = sorted(proc_totals.items(), key=lambda x: x[0])

    grand_qtd = 0
    grand_val = 0.0
    row_idx   = 7

    for i, (proc, qtd) in enumerate(sorted_procs):
        fill  = s["fill_even"] if i % 2 == 0 else s["fill_odd"]
        pinfo = proc_data.get(proc, {"name": "", "value": 0.0})
        val   = pinfo["value"]
        total = qtd * val
        grand_qtd += qtd
        grand_val += total

        ws[f"A{row_idx}"].value = f"  {proc}"
        ws[f"B{row_idx}"].value = f"  {pinfo['name']}"
        ws[f"C{row_idx}"].value = qtd
        ws[f"D{row_idx}"].value = val
        ws[f"D{row_idx}"].number_format = 'R$ #,##0.00'
        ws[f"E{row_idx}"].value = total
        ws[f"E{row_idx}"].number_format = 'R$ #,##0.00'

        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row_idx}"].font      = s["font_proc"]
            ws[f"{col}{row_idx}"].fill      = fill
            ws[f"{col}{row_idx}"].border    = s["border"]
            ws[f"{col}{row_idx}"].alignment = (
                s["al_center"] if col in ["A", "C"] else
                s["al_right"]  if col in ["D", "E"] else s["al_left"])
        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

    # Linha TOTAL GERAL
    fill_geral = PatternFill("solid", start_color="0D2E4E")
    font_geral = Font(name="Arial", bold=True, size=11, color="FFFFFF")

    ws.merge_cells(f"A{row_idx}:B{row_idx}")
    for col, val, aln, fmt in [
        ("A", "TOTAL GERAL", s["al_right"],  None),
        ("C", grand_qtd,    s["al_center"], None),
        ("D", "",           s["al_right"],  None),
        ("E", grand_val,    s["al_right"],  'R$ #,##0.00'),
    ]:
        ws[f"{col}{row_idx}"].value     = val
        ws[f"{col}{row_idx}"].font      = font_geral
        ws[f"{col}{row_idx}"].fill      = fill_geral
        ws[f"{col}{row_idx}"].alignment = aln
        ws[f"{col}{row_idx}"].border    = s["border"]
        if fmt: ws[f"{col}{row_idx}"].number_format = fmt
    ws.row_dimensions[row_idx].height = 20

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18


# ─────────────────────────────────────────────────────────────
# 10. APAC — Leitura e geração de relatório
# ─────────────────────────────────────────────────────────────

EXTENSOES_APAC = {
    'JAN','FEV','MAR','ABR','MAI','JUN',
    'JUL','AGO','SET','OUT','NOV','DEZ','TXT'
}

def detectar_tipo_arquivo(caminho):
    """Retorna 'BPA', 'APAC' ou 'DESCONHECIDO' baseado no conteúdo."""
    with open(caminho, encoding='ISO-8859-1') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith('01#BPA#'):
                return 'BPA'
            if line.startswith('01#APAC'):
                return 'APAC'
            break
    return 'DESCONHECIDO'


def ler_arquivo_apac(caminho):
    """Lê arquivo APAC e retorna info + dados estruturados."""
    info       = {}
    apac_corpo = {}
    apac_procs = collections.defaultdict(list)

    with open(caminho, encoding='ISO-8859-1') as f:
        for line in f:
            line = line.rstrip('\r\n')
            tipo = line[:2]

            # Cabeçalho
            if tipo == '01' and '#APAC' in line[:10]:
                comp_raw       = line[7:13]
                info['comp']   = f"{comp_raw[4:6]}/{comp_raw[0:4]}"
                info['nome']   = line[23:53].strip()
                info['sigla']  = line[53:59].strip()
                info['cnpj']   = line[59:73].strip()
                info['dest']   = line[73:113].strip()

            # Corpo da APAC (14)
            elif tipo == '14' and len(line) >= 230:
                num           = line[8:21]
                cnes          = line[23:30].strip()
                nome_pac      = line[57:87].strip()
                nasc_raw      = line[177:185].strip()
                nasc          = (f"{nasc_raw[6:8]}/{nasc_raw[4:6]}/{nasc_raw[0:4]}"
                                 if len(nasc_raw) == 8 else '')
                sexo          = line[185] if len(line) > 185 else ''
                proc_princ    = line[216:226].strip()
                mot_saida     = line[226:228].strip()
                nome_med      = line[236:266].strip()
                cid_ca        = line[311:315].strip() if len(line) > 315 else ''
                if not info.get('cnes'):
                    info['cnes'] = cnes
                apac_corpo[num] = {
                    'nome':       nome_pac,
                    'nasc':       nasc,
                    'sexo':       sexo,
                    'proc_princ': proc_princ,
                    'mot_saida':  mot_saida,
                    'medico':     nome_med,
                    'cid':        cid_ca,
                }

            # Procedimentos da APAC (13)
            elif tipo == '13' and len(line) >= 68:
                num   = line[8:21]
                proc  = line[21:31].strip()
                cbo   = line[31:37].strip()
                try:
                    qtd = int(line[37:44].strip())
                except ValueError:
                    qtd = 0
                cid_p = line[64:68].strip()
                apac_procs[num].append((proc, cbo, qtd, cid_p))

    return info, apac_corpo, apac_procs


def _cabecalho_apac(ws, titulo, info, ncols=4):
    s    = _estilos()
    last = chr(ord('A') + ncols - 1)

    ws.merge_cells(f"A1:{last}1")
    ws["A1"].value     = titulo
    ws["A1"].font      = s["font_title"]
    ws["A1"].fill      = s["fill_hdr_bg"]
    ws["A1"].alignment = s["al_center"]
    ws.row_dimensions[1].height = 26

    cnes_cnpj = info.get('cnes', '') + '  /  CNPJ: ' + info.get('cnpj', '')
    for rn, lbl, val in [
        (2, 'CNES / CNPJ',            cnes_cnpj),
        (3, 'NOME DO ESTABELECIMENTO', info.get('nome', '')),
        (4, 'COMPETÊNCIA',             info.get('comp', '')),
    ]:
        ws[f"A{rn}"].value     = lbl
        ws[f"A{rn}"].font      = s["font_lbl"]
        ws[f"A{rn}"].fill      = s["fill_hdr_lbl"]
        ws[f"A{rn}"].alignment = s["al_center"]
        ws.merge_cells(f"B{rn}:{last}{rn}")
        ws[f"B{rn}"].value     = val
        ws[f"B{rn}"].font      = s["font_val"]
        ws[f"B{rn}"].fill      = s["fill_hdr_val"]
        ws[f"B{rn}"].alignment = s["al_left"]
        ws.row_dimensions[rn].height = 18

    ws.merge_cells(f"A5:{last}5")
    ws["A5"].fill = s["fill_hdr_bg"]
    ws.row_dimensions[5].height = 6


def build_sheet_apac_resumo(ws, info, apac_procs, proc_data):
    """Aba Resumo Geral APAC: Cód → Descrição → Qtd → Valor."""
    s = _estilos()
    _cabecalho_apac(ws, "RESUMO GERAL — APAC", info, ncols=4)

    _cabecalho_apac(ws, "RESUMO GERAL — APAC", info, ncols=5)
    for col, title in [
        ("A6", "Cód. Procedimento"),
        ("B6", "Descrição do Procedimento"),
        ("C6", "Quantidade"),
        ("D6", "Valor Unit. (R$)"),
        ("E6", "Valor Total (R$)"),
    ]:
        ws[col].value     = title
        ws[col].font      = s["font_col_hdr"]
        ws[col].fill      = s["fill_col_hdr"]
        ws[col].alignment = s["al_center"]
        ws[col].border    = s["border"]
    ws.row_dimensions[6].height = 20
    ws.freeze_panes = "A7"

    proc_totals = collections.defaultdict(int)
    for procs in apac_procs.values():
        for proc, _, qtd, _ in procs:
            proc_totals[proc] += qtd

    grand_qtd = 0; grand_val = 0.0
    row_idx = 7
    for i, (proc, qtd) in enumerate(sorted(proc_totals.items())):
        pinfo = proc_data.get(proc, {"name": "", "value": 0.0})
        fill  = s["fill_even"] if i % 2 == 0 else s["fill_odd"]
        total = qtd * pinfo["value"]
        grand_qtd += qtd; grand_val += total

        ws[f"A{row_idx}"].value = f"  {proc}"
        ws[f"B{row_idx}"].value = f"  {pinfo['name']}"
        ws[f"C{row_idx}"].value = qtd
        ws[f"D{row_idx}"].value = pinfo["value"]
        ws[f"D{row_idx}"].number_format = 'R$ #,##0.00'
        ws[f"E{row_idx}"].value = total
        ws[f"E{row_idx}"].number_format = 'R$ #,##0.00'
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row_idx}"].font      = s["font_proc"]
            ws[f"{col}{row_idx}"].fill      = fill
            ws[f"{col}{row_idx}"].border    = s["border"]
            ws[f"{col}{row_idx}"].alignment = (
                s["al_center"] if col in ["A", "C"] else
                s["al_right"]  if col in ["D", "E"] else s["al_left"])
        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

    # Total Geral
    fill_g = PatternFill("solid", start_color="0D2E4E")
    font_g = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.merge_cells(f"A{row_idx}:B{row_idx}")
    for col, val, aln, fmt in [
        ("A", "TOTAL GERAL", s["al_right"],  None),
        ("C", grand_qtd,    s["al_center"], None),
        ("D", "",           s["al_right"],  None),
        ("E", grand_val,    s["al_right"],  'R$ #,##0.00'),
    ]:
        ws[f"{col}{row_idx}"].value     = val
        ws[f"{col}{row_idx}"].font      = font_g
        ws[f"{col}{row_idx}"].fill      = fill_g
        ws[f"{col}{row_idx}"].alignment = aln
        ws[f"{col}{row_idx}"].border    = s["border"]
        if fmt: ws[f"{col}{row_idx}"].number_format = fmt
    ws.row_dimensions[row_idx].height = 20

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18


def build_sheet_apac_pacientes(ws, info, apac_corpo, apac_procs, proc_data):
    """Aba Por Paciente APAC: nome → procedimentos → total."""
    s = _estilos()
    _cabecalho_apac(ws, "APAC — RELATÓRIO POR PACIENTE", info, ncols=9)

    headers = [
        ("A6", "Nº APAC"), ("B6", "Nome do Paciente"), ("C6", "Nasc."),
        ("D6", "Sexo"),    ("E6", "Cód. Procedimento"), ("F6", "Descrição"),
        ("G6", "Qtd"),     ("H6", "Valor Unit. (R$)"),  ("I6", "CID"),
    ]
    for col, title in headers:
        ws[col].value     = title
        ws[col].font      = s["font_col_hdr"]
        ws[col].fill      = s["fill_col_hdr"]
        ws[col].alignment = s["al_center"]
        ws[col].border    = s["border"]
    ws.row_dimensions[6].height = 20
    ws.freeze_panes = "A7"

    row_idx = 7
    for num in sorted(apac_corpo.keys()):
        corpo = apac_corpo[num]
        procs = apac_procs.get(num, [])
        tq = sum(q for _, _, q, _ in procs)
        tv = sum(q * proc_data.get(p, {"value": 0})["value"] for p, _, q, _ in procs)
        sexo_lbl = "Masc." if corpo["sexo"] == "M" else "Fem." if corpo["sexo"] == "F" else ""

        # Linha cabeçalho do paciente
        ws.merge_cells(f"A{row_idx}:I{row_idx}")
        ws[f"A{row_idx}"].value = (
            f"  {corpo['nome']}  |  {sexo_lbl}  Nasc: {corpo['nasc']}"
            f"  |  APAC: {num}  |  Médico: {corpo['medico']}")
        ws[f"A{row_idx}"].font      = s["font_name"]
        ws[f"A{row_idx}"].fill      = s["fill_name"]
        ws[f"A{row_idx}"].alignment = s["al_left"]
        ws[f"A{row_idx}"].border    = s["border"]
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

        for i, (proc, cbo, qtd, cid) in enumerate(procs):
            fill  = s["fill_even"] if i % 2 == 0 else s["fill_odd"]
            pinfo = proc_data.get(proc, {"name": "", "value": 0.0})
            row_data = [
                ("A", f"  {num}",          s["al_center"]),
                ("B", f"  {corpo['nome']}", s["al_left"]),
                ("C", corpo["nasc"],        s["al_center"]),
                ("D", corpo["sexo"],        s["al_center"]),
                ("E", f"  {proc}",          s["al_center"]),
                ("F", f"  {pinfo['name']}", s["al_left"]),
                ("G", qtd,                  s["al_center"]),
                ("H", pinfo["value"],       s["al_right"]),
                ("I", cid,                  s["al_center"]),
            ]
            for col, val, aln in row_data:
                ws[f"{col}{row_idx}"].value     = val
                ws[f"{col}{row_idx}"].font      = s["font_proc"]
                ws[f"{col}{row_idx}"].fill      = fill
                ws[f"{col}{row_idx}"].alignment = aln
                ws[f"{col}{row_idx}"].border    = s["border"]
            ws[f"H{row_idx}"].number_format = 'R$ #,##0.00'
            ws.row_dimensions[row_idx].height = 16
            row_idx += 1

        # Total do paciente
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        ws[f"A{row_idx}"].value     = "TOTAL DO PACIENTE"
        ws[f"A{row_idx}"].font      = s["font_total"]
        ws[f"A{row_idx}"].fill      = s["fill_total"]
        ws[f"A{row_idx}"].alignment = s["al_right"]
        ws[f"A{row_idx}"].border    = s["border"]
        for col, val, aln in [("G", tq, s["al_center"]), ("H", tv, s["al_right"])]:
            ws[f"{col}{row_idx}"].value     = val
            ws[f"{col}{row_idx}"].font      = s["font_total"]
            ws[f"{col}{row_idx}"].fill      = s["fill_total"]
            ws[f"{col}{row_idx}"].alignment = aln
            ws[f"{col}{row_idx}"].border    = s["border"]
        ws[f"H{row_idx}"].number_format = 'R$ #,##0.00'
        ws.row_dimensions[row_idx].height = 17
        row_idx += 1

    for col, w in zip("ABCDEFGHI", [16, 30, 12, 8, 16, 42, 8, 16, 10]):
        ws.column_dimensions[col].width = w


def gerar_excel_apac(caminho, caminho_saida=None, callback=None):
    """Gera relatório Excel para arquivo APAC."""
    if callback:
        callback("Carregando tabela SIGTAP…")
    sigtap_zip = baixar_sigtap(callback)
    proc_data  = carregar_procedimentos(sigtap_zip)

    if callback:
        callback("Lendo arquivo APAC…")
    info, apac_corpo, apac_procs = ler_arquivo_apac(caminho)

    n_apac = len(apac_corpo)
    if callback:
        callback(f"Gerando relatório APAC — {n_apac} APACs…")

    wb = Workbook()

    # Aba 1: Resumo Geral
    ws1 = wb.active
    ws1.title = "Resumo Geral"
    build_sheet_apac_resumo(ws1, info, apac_procs, proc_data)

    # Aba 2: Por Paciente
    ws2 = wb.create_sheet("Por Paciente")
    build_sheet_apac_pacientes(ws2, info, apac_corpo, apac_procs, proc_data)

    wb.active = ws1

    if caminho_saida is None:
        base = os.path.splitext(caminho)[0]
        caminho_saida = base + "_relatorio_apac.xlsx"

    wb.save(caminho_saida)
    return caminho_saida, n_apac


# ─────────────────────────────────────────────────────────────
# 11. Conferência PA vs PDF de Faturamento
# ─────────────────────────────────────────────────────────────

import unicodedata as _ud
import re as _re

def _norm(s):
    s = _ud.normalize('NFD', s)
    s = ''.join(c for c in s if _ud.category(c) != 'Mn')
    return _re.sub(r'\s+', ' ', s.upper().strip())


def ler_pdf_faturamento(caminho_pdf, callback=None):
    """
    Extrai dados do PDF de faturamento.
    Tenta pdftotext (poppler) primeiro, depois pdfplumber.
    Instala pdfplumber automaticamente se necessário.
    Retorna dict: {nome_normalizado: {codigo: qtd}}
    """
    lines = None

    # Tentativa 1: pdftotext (poppler-utils)
    try:
        import subprocess, tempfile
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tmp:
            tmp_path = tmp.name
        result = subprocess.run(
            ['pdftotext', '-layout', caminho_pdf, tmp_path],
            capture_output=True, timeout=120)
        if result.returncode == 0:
            with open(tmp_path, encoding='utf-8', errors='replace') as f:
                lines = f.readlines()
            os.unlink(tmp_path)
    except Exception:
        pass

    # Tentativa 2: pdfplumber
    if lines is None:
        try:
            import pdfplumber
        except ImportError:
            if callback: callback("Instalando pdfplumber…")
            import subprocess as _sp
            _sp.run([sys.executable, '-m', 'pip', 'install', 'pdfplumber', '--quiet'],
                    capture_output=True)
            try:
                import pdfplumber
            except ImportError:
                raise ImportError(
                    "Nao foi possivel instalar pdfplumber.\n"
                    "Execute manualmente: pip install pdfplumber")

        lines = []
        with pdfplumber.open(caminho_pdf) as pdf:
            total = len(pdf.pages)
            if callback: callback(f"Lendo PDF ({total} paginas)…")
            for i, page in enumerate(pdf.pages, 1):
                if callback and i % 50 == 0:
                    callback(f"Lendo PDF… pagina {i}/{total}")
                txt = page.extract_text()
                if txt:
                    lines += txt.split('\n')

    if callback: callback(f"Analisando {len(lines)} linhas do PDF…")

    pdf_data = collections.defaultdict(lambda: collections.defaultdict(int))
    current_name = None

    for line in lines:
        # Linha de paciente — dois formatos:
        # pdftotext:   "001.004.00006792   04/03/2026 ROBERTA KOSSAK..."  (espaços separando)
        # pdfplumber:  "001.004.00006792 04/03/2026ROBERTA KOSSAK..."    (data colada ao nome)
        m = _re.match(
            r'\s*\d{3}\.\d{3}\.\d{8}\s+\d{2}/\d{2}/\d{4}\s*([A-Z].+?)(?:\s{3,}|$)',
            line)
        if m:
            # Extrai o nome: pode estar colado à data ou separado por espaços
            raw = m.group(1).strip()
            # Remove tudo após 3+ espaços (carteirinha, CNS etc)
            nome = _re.split(r'\s{3,}', raw)[0].strip()
            current_name = _norm(nome)
            continue

        # Linha de exame — formatos:
        # pdftotext:   "   0 202010317   CREATININA   1  ..."
        # pdfplumber:  "0202010317 CREATININA 1 1,85 0,00 1,85"
        #              "00202031110 VDRL 1 2,83 ..."  (código com 11 dígitos: 2 zeros extras)
        m2 = _re.match(
            r'\s*(?:0\s+)?(0*\d{9,10})\s+([A-ZÁÉÍÓÚÀÃÂÊÎÔÛÇ])', line)
        if m2 and current_name:
            raw_code = m2.group(1).lstrip('0')
            # Garante 10 dígitos
            code = raw_code.zfill(10) if len(raw_code) <= 10 else raw_code[-10:]
            pdf_data[current_name][code] += 1

    return pdf_data


def conferir_pa_vs_pdf(caminho_pa, caminho_pdf, caminho_saida=None, callback=None):
    """
    Compara arquivo PA/MAI com PDF de faturamento.
    Gera Excel com resumo e divergências por paciente.
    """
    if callback: callback("Carregando tabela SIGTAP…")
    sigtap_zip = baixar_sigtap(callback)
    proc_data  = carregar_procedimentos(sigtap_zip)

    if callback: callback("Lendo arquivo PA/MAI…")
    # Reutiliza ler_arquivo (BPA-I)
    info_bpa, bpa_i, _, _, _, _ = ler_arquivo(caminho_pa)
    # Flatten bpa_i: {nome: {proc: qtd_total}}
    mai_data = collections.defaultdict(lambda: collections.defaultdict(int))
    for nome, datas in bpa_i.items():
        for dp in datas.values():
            for proc, qtd in dp.items():
                mai_data[_norm(nome)][proc] += qtd

    if callback: callback("Lendo PDF de faturamento…")
    pdf_data = ler_pdf_faturamento(caminho_pdf, callback)

    if callback: callback("Cruzando dados…")

    # Match de nomes (MAI trunca em 30 chars)
    pdf_by_prefix = {}
    for pn in pdf_data:
        if pn[:30] not in pdf_by_prefix:
            pdf_by_prefix[pn[:30]] = pn

    mai_to_pdf = {}
    for mn in mai_data:
        if mn in pdf_data:
            mai_to_pdf[mn] = mn
        elif mn in pdf_by_prefix:
            mai_to_pdf[mn] = pdf_by_prefix[mn]
        else:
            for pn in pdf_data:
                if pn.startswith(mn) or mn.startswith(pn[:len(mn)]):
                    mai_to_pdf[mn] = pn
                    break

    matched  = set(mai_to_pdf)
    only_mai = set(mai_data) - matched
    only_pdf = set(pdf_data) - set(mai_to_pdf.values())

    cor_tipo = {
        'QTD DIFERENTE':           'FFF3CD',
        'SOMENTE NO PDF':          'D4EDDA',
        'SOMENTE NO MAI':          'F8D7DA',
        'PACIENTE SOMENTE NO PDF': 'C3E6CB',
        'PACIENTE SOMENTE NO MAI': 'F5C6CB',
    }

    divergencias = []

    for mn in sorted(matched):
        pn = mai_to_pdf[mn]
        pp = pdf_data[pn]; pm = mai_data[mn]
        for proc in sorted(set(pp) | set(pm)):
            qp = pp.get(proc, 0); qm = pm.get(proc, 0)
            if qp != qm:
                divergencias.append({
                    'nome': mn, 'proc': proc,
                    'desc': proc_data.get(proc, {"name":""})["name"],
                    'qtd_pdf': qp, 'qtd_mai': qm,
                    'tipo': 'QTD DIFERENTE'   if qp>0 and qm>0
                            else('SOMENTE NO PDF' if qm==0 else 'SOMENTE NO MAI')})

    for mn in sorted(only_mai):
        for proc, qtd in sorted(mai_data[mn].items()):
            divergencias.append({
                'nome': mn, 'proc': proc,
                'desc': proc_data.get(proc, {"name":""})["name"],
                'qtd_pdf': 0, 'qtd_mai': qtd, 'tipo': 'PACIENTE SOMENTE NO MAI'})

    for pn in sorted(only_pdf):
        for proc, qtd in sorted(pdf_data[pn].items()):
            divergencias.append({
                'nome': pn, 'proc': proc,
                'desc': proc_data.get(proc, {"name":""})["name"],
                'qtd_pdf': qtd, 'qtd_mai': 0, 'tipo': 'PACIENTE SOMENTE NO PDF'})

    from collections import Counter
    tipos      = Counter(d['tipo'] for d in divergencias)
    n_pac_div  = len(set(d['nome'] for d in divergencias))
    n_qtd_dif  = sum(1 for d in divergencias if d['tipo'] == 'QTD DIFERENTE')

    if callback: callback(f"Gerando Excel — {len(divergencias)} divergências…")

    # ── Excel ────────────────────────────────────────────
    s = _estilos()
    wb = Workbook()

    def _hdr_cell(ws, col, row, val, font, fill, aln, brd=None):
        c = ws[f"{col}{row}"]
        c.value = val; c.font = font; c.fill = fill; c.alignment = aln
        if brd: c.border = brd

    # ABA 1: Resumo
    ws1 = wb.active; ws1.title = 'Resumo'
    ws1.merge_cells('A1:D1')
    ws1['A1'].value = 'CONFERÊNCIA DE FATURAMENTO — PA vs PDF'
    ws1['A1'].font = s["font_title"]; ws1['A1'].fill = s["fill_hdr_bg"]
    ws1['A1'].alignment = s["al_center"]; ws1.row_dimensions[1].height = 26

    nome_est = info_bpa.get('nome_cnes', '') or info_bpa.get('nome', '')
    for rn, lbl, val in [
        (2, 'ESTABELECIMENTO', nome_est),
        (3, 'CNES',            info_bpa.get('cnes', '')),
        (4, 'COMPETÊNCIA',     info_bpa.get('competencia', '')),
        (5, 'ARQUIVO PA',      os.path.basename(caminho_pa)),
        (6, 'ARQUIVO PDF',     os.path.basename(caminho_pdf)),
    ]:
        ws1[f'A{rn}'].value = lbl; ws1[f'A{rn}'].font = s["font_lbl"]
        ws1[f'A{rn}'].fill = s["fill_hdr_lbl"]; ws1[f'A{rn}'].alignment = s["al_center"]
        ws1.merge_cells(f'B{rn}:D{rn}')
        ws1[f'B{rn}'].value = val; ws1[f'B{rn}'].font = s["font_val"]
        ws1[f'B{rn}'].fill = s["fill_hdr_val"]; ws1[f'B{rn}'].alignment = s["al_left"]
        ws1.row_dimensions[rn].height = 18

    ws1.merge_cells('A7:D7'); ws1['A7'].fill = s["fill_hdr_bg"]; ws1.row_dimensions[7].height = 8

    for col, title in [('A','Indicador'),('B','Pacientes'),('C','Procedimentos'),('D','')]:
        ws1[f'{col}8'].value = title; ws1[f'{col}8'].font = s["font_col_hdr"]
        ws1[f'{col}8'].fill = s["fill_col_hdr"]; ws1[f'{col}8'].alignment = s["al_center"]
        ws1[f'{col}8'].border = s["border"]
    ws1.row_dimensions[8].height = 20

    stats = [
        ('Total pacientes no PDF',      len(pdf_data), sum(len(v) for v in pdf_data.values()),  'DEEAF1'),
        ('Total pacientes no PA/MAI',   len(mai_data), sum(len(v) for v in mai_data.values()),  'DEEAF1'),
        ('Pacientes correspondidos',    len(matched),  '',                                       'D4EDDA'),
        ('Pacientes só no PDF',         len(only_pdf), sum(len(pdf_data[n]) for n in only_pdf), 'FFF3CD'),
        ('Pacientes só no PA/MAI',      len(only_mai), sum(len(mai_data[n]) for n in only_mai), 'F8D7DA'),
        ('⚠️  Com divergência de QTD',  n_qtd_dif,    n_qtd_dif,                               'FFE0B2'),
        ('Total divergências',          n_pac_div,    len(divergencias),                        'FFE0B2'),
    ]
    for i, (lbl, pac, proc, cor) in enumerate(stats):
        r = 9 + i; fx = PatternFill("solid", start_color=cor)
        for col, val, aln in [('A',lbl,s["al_left"]),('B',pac,s["al_center"]),
                               ('C',proc,s["al_center"]),('D','',s["al_left"])]:
            ws1[f'{col}{r}'].value = val; ws1[f'{col}{r}'].font = s["font_proc"]
            ws1[f'{col}{r}'].fill = fx; ws1[f'{col}{r}'].alignment = aln
            ws1[f'{col}{r}'].border = s["border"]
        ws1.row_dimensions[r].height = 16

    ws1.merge_cells('A16:D16'); ws1['A16'].fill = s["fill_hdr_bg"]; ws1.row_dimensions[16].height = 8
    ws1.merge_cells('A17:D17'); ws1['A17'].value = 'DIVERGÊNCIAS POR TIPO'
    ws1['A17'].font = s["font_col_hdr"]; ws1['A17'].fill = s["fill_col_hdr"]
    ws1['A17'].alignment = s["al_center"]; ws1.row_dimensions[17].height = 20

    for col, title in [('A','Tipo de Divergência'),('B','Qtd'),('C',''),('D','')]:
        ws1[f'{col}18'].value = title; ws1[f'{col}18'].font = s["font_col_hdr"]
        ws1[f'{col}18'].fill = s["fill_hdr_val"]; ws1[f'{col}18'].alignment = s["al_center"]
        ws1[f'{col}18'].border = s["border"]
    ws1.row_dimensions[18].height = 18

    r = 19
    for tipo, cnt in tipos.most_common():
        fx = PatternFill("solid", start_color=cor_tipo.get(tipo, 'FFFFFF'))
        ws1.merge_cells(f'A{r}:C{r}')
        ws1[f'A{r}'].value = f'  {tipo}'; ws1[f'A{r}'].font = s["font_proc"]
        ws1[f'A{r}'].fill = fx; ws1[f'A{r}'].alignment = s["al_left"]; ws1[f'A{r}'].border = s["border"]
        ws1[f'D{r}'].value = cnt; ws1[f'D{r}'].font = Font(name="Arial", bold=True, size=10)
        ws1[f'D{r}'].fill = fx; ws1[f'D{r}'].alignment = s["al_center"]; ws1[f'D{r}'].border = s["border"]
        ws1.row_dimensions[r].height = 16; r += 1

    for col, w in zip('ABCD', [40, 14, 16, 14]): ws1.column_dimensions[col].width = w

    # ABA 2: Divergências por Paciente
    ws2 = wb.create_sheet('Divergências por Paciente')
    ws2.merge_cells('A1:G1')
    ws2['A1'].value = 'DIVERGÊNCIAS — DETALHE POR PACIENTE'
    ws2['A1'].font = s["font_title"]; ws2['A1'].fill = s["fill_hdr_bg"]
    ws2['A1'].alignment = s["al_center"]; ws2.row_dimensions[1].height = 26
    ws2.merge_cells('A2:G2'); ws2['A2'].fill = s["fill_hdr_bg"]; ws2.row_dimensions[2].height = 6

    for col, title in [('A','Nome do Paciente'),('B','Cód. Proc.'),
                       ('C','Descrição do Procedimento'),('D','Qtd PDF'),
                       ('E','Qtd PA/MAI'),('F','Dif.'),('G','Tipo')]:
        ws2[f'{col}3'].value = title; ws2[f'{col}3'].font = s["font_col_hdr"]
        ws2[f'{col}3'].fill = s["fill_col_hdr"]; ws2[f'{col}3'].alignment = s["al_center"]
        ws2[f'{col}3'].border = s["border"]
    ws2.row_dimensions[3].height = 20; ws2.freeze_panes = 'A4'

    r = 4; nome_ant = None
    for d in divergencias:
        if d['nome'] != nome_ant:
            nome_ant = d['nome']
            ws2.merge_cells(f'A{r}:G{r}')
            ws2[f'A{r}'].value = f"  {d['nome']}"
            ws2[f'A{r}'].font = s["font_name"]; ws2[f'A{r}'].fill = s["fill_name"]
            ws2[f'A{r}'].alignment = s["al_left"]; ws2[f'A{r}'].border = s["border"]
            ws2.row_dimensions[r].height = 18; r += 1

        fx  = PatternFill("solid", start_color=cor_tipo.get(d['tipo'], 'FFFFFF'))
        dif = d['qtd_mai'] - d['qtd_pdf']
        for col, val, aln in [
            ('A', f"  {d['nome']}",  s["al_left"]),
            ('B', d['proc'],         s["al_center"]),
            ('C', d['desc'],         s["al_left"]),
            ('D', d['qtd_pdf'],      s["al_center"]),
            ('E', d['qtd_mai'],      s["al_center"]),
            ('F', f"+{dif}" if dif > 0 else str(dif), s["al_center"]),
            ('G', d['tipo'],         s["al_center"]),
        ]:
            ws2[f'{col}{r}'].value = val; ws2[f'{col}{r}'].font = s["font_proc"]
            ws2[f'{col}{r}'].fill = fx; ws2[f'{col}{r}'].alignment = aln
            ws2[f'{col}{r}'].border = s["border"]
        ws2.row_dimensions[r].height = 16; r += 1

    for col, w in zip('ABCDEFG', [32, 14, 44, 10, 12, 8, 28]):
        ws2.column_dimensions[col].width = w

    # ABA 3: Consolidado por Paciente
    ws3 = wb.create_sheet('Pacientes com Divergência')
    ws3.merge_cells('A1:F1')
    ws3['A1'].value = 'PACIENTES COM DIVERGÊNCIA — CONSOLIDADO'
    ws3['A1'].font = s["font_title"]; ws3['A1'].fill = s["fill_hdr_bg"]
    ws3['A1'].alignment = s["al_center"]; ws3.row_dimensions[1].height = 26
    ws3.merge_cells('A2:F2'); ws3['A2'].fill = s["fill_hdr_bg"]; ws3.row_dimensions[2].height = 6

    for col, title in [
        ('A', 'Nome do Paciente'),
        ('B', 'QTD Diferente'),
        ('C', 'Só no PDF'),
        ('D', 'Só no PA/MAI'),
        ('E', 'Só no PDF (pac.)'),
        ('F', 'Total Divergências'),
    ]:
        ws3[f'{col}3'].value = title
        ws3[f'{col}3'].font = s["font_col_hdr"]
        ws3[f'{col}3'].fill = s["fill_col_hdr"]
        ws3[f'{col}3'].alignment = s["al_center"]
        ws3[f'{col}3'].border = s["border"]
    ws3.row_dimensions[3].height = 20
    ws3.freeze_panes = 'A4'

    # Consolidar divergências por paciente
    pac_div = collections.defaultdict(lambda: {
        'QTD DIFERENTE': 0, 'SOMENTE NO PDF': 0,
        'SOMENTE NO MAI': 0, 'PACIENTE SOMENTE NO PDF': 0,
        'PACIENTE SOMENTE NO MAI': 0,
    })
    for d in divergencias:
        pac_div[d['nome']][d['tipo']] += 1

    r = 4
    fill_tot = PatternFill("solid", start_color="0D2E4E")
    font_tot = Font(name="Arial", bold=True, size=11, color="FFFFFF")

    grand_qtd_dif = grand_so_pdf = grand_so_mai = grand_so_pdf_pac = grand_total = 0

    for i, nome in enumerate(sorted(pac_div.keys())):
        d = pac_div[nome]
        qtd_dif    = d['QTD DIFERENTE']
        so_pdf     = d['SOMENTE NO PDF']
        so_mai     = d['SOMENTE NO MAI']
        so_pdf_pac = d['PACIENTE SOMENTE NO PDF']
        so_mai_pac = d['PACIENTE SOMENTE NO MAI']
        total      = qtd_dif + so_pdf + so_mai + so_pdf_pac + so_mai_pac

        grand_qtd_dif    += qtd_dif
        grand_so_pdf     += so_pdf
        grand_so_mai     += so_mai
        grand_so_pdf_pac += so_pdf_pac
        grand_total      += total

        # Cor da linha
        if qtd_dif > 0:
            fx = PatternFill("solid", start_color="FFF3CD")   # amarelo
        elif so_mai_pac > 0:
            fx = PatternFill("solid", start_color="F8D7DA")   # vermelho claro
        elif so_pdf_pac > 0:
            fx = PatternFill("solid", start_color="C3E6CB")   # verde claro
        else:
            fx = s["fill_even"] if i % 2 == 0 else s["fill_odd"]

        for col, val, aln in [
            ('A', f'  {nome}',  s["al_left"]),
            ('B', qtd_dif,      s["al_center"]),
            ('C', so_pdf,       s["al_center"]),
            ('D', so_mai,       s["al_center"]),
            ('E', so_pdf_pac,   s["al_center"]),
            ('F', total,        s["al_center"]),
        ]:
            ws3[f'{col}{r}'].value     = val
            ws3[f'{col}{r}'].font      = s["font_proc"]
            ws3[f'{col}{r}'].fill      = fx
            ws3[f'{col}{r}'].alignment = aln
            ws3[f'{col}{r}'].border    = s["border"]
        ws3.row_dimensions[r].height = 16
        r += 1

    # Linha TOTAL
    ws3[f'A{r}'].value = 'TOTAL GERAL'
    ws3[f'A{r}'].font = font_tot; ws3[f'A{r}'].fill = fill_tot
    ws3[f'A{r}'].alignment = s["al_right"]; ws3[f'A{r}'].border = s["border"]
    for col, val in [('B', grand_qtd_dif), ('C', grand_so_pdf),
                     ('D', grand_so_mai),  ('E', grand_so_pdf_pac), ('F', grand_total)]:
        ws3[f'{col}{r}'].value     = val
        ws3[f'{col}{r}'].font      = font_tot
        ws3[f'{col}{r}'].fill      = fill_tot
        ws3[f'{col}{r}'].alignment = s["al_center"]
        ws3[f'{col}{r}'].border    = s["border"]
    ws3.row_dimensions[r].height = 20

    ws3.column_dimensions['A'].width = 36
    for col in 'BCDEF': ws3.column_dimensions[col].width = 16

    wb.active = ws1
    if caminho_saida is None:
        base = os.path.splitext(caminho_pa)[0]
        caminho_saida = base + '_conferencia.xlsx'
    wb.save(caminho_saida)
    return caminho_saida, len(divergencias), n_pac_div, tipos
